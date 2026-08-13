"""
mock_data_reader.py — read the Phase-2 ``mock_data/`` tree.

The harness's mock API services live under ``mock_data/<slug>-api/`` as
CSV / XLSX / JSON / JSONL / TXT files. The generator reads them for two
purposes:

  1. Verify that every Value Lock concrete value (from
     ``golden_steer_flow.md`` Section 3) actually appears somewhere in
     the active-service files — i.e. the agent has a chance to find it.
  2. Surface real entity names + values to the LLM rubric prompt so
     criteria embed concrete identifiers (Rule 3, self-contained).

XLSX support uses ``openpyxl`` when installed; if missing, ``.xlsx``
files are skipped with a warning rather than failing.
"""
from __future__ import annotations

import csv
import io
import json
import re
from dataclasses import dataclass, field
from pathlib import Path


SUPPORTED_SUFFIXES = {".csv", ".tsv", ".json", ".jsonl", ".txt", ".md", ".xlsx"}
MAX_TEXT_BYTES = 2_000_000


class MockDataError(ValueError):
    """Raised when the mock_data tree is missing or unreadable."""


@dataclass
class MockFile:
    relative_path: str
    suffix: str
    rows: list[dict] = field(default_factory=list)
    raw_text: str = ""
    parse_error: str | None = None

    @property
    def service(self) -> str:
        parts = self.relative_path.split("/", 1)
        return parts[0] if parts else ""

    @property
    def filename(self) -> str:
        return self.relative_path.split("/")[-1]


@dataclass
class MockData:
    root: Path
    files: list[MockFile] = field(default_factory=list)

    def by_service(self) -> dict[str, list[MockFile]]:
        out: dict[str, list[MockFile]] = {}
        for f in self.files:
            out.setdefault(f.service, []).append(f)
        return out

    def services(self) -> list[str]:
        return sorted({f.service for f in self.files})

    def file_count(self) -> int:
        return len(self.files)


def load(mock_data_root: Path) -> MockData:
    """Walk ``mock_data_root`` recursively and parse every supported file.

    Raises ``MockDataError`` if the directory is missing or empty.
    """
    if not mock_data_root.exists():
        raise MockDataError(
            f"mock_data directory not found at {mock_data_root}. The v2 "
            "generator requires the Kensei Phase-2 mock_data tree (one "
            "{slug}-api/ subdirectory per active service)."
        )
    if not mock_data_root.is_dir():
        raise MockDataError(f"mock_data path is not a directory: {mock_data_root}")

    data = MockData(root=mock_data_root)
    for path in sorted(mock_data_root.rglob("*")):
        if not path.is_file():
            continue
        suffix = path.suffix.lower()
        if suffix not in SUPPORTED_SUFFIXES:
            continue
        rel = path.relative_to(mock_data_root).as_posix()
        mock_file = MockFile(relative_path=rel, suffix=suffix)
        try:
            _parse_into(path, mock_file)
        except Exception as exc:
            mock_file.parse_error = f"{type(exc).__name__}: {exc}"
        data.files.append(mock_file)

    if not data.files:
        raise MockDataError(
            f"mock_data directory at {mock_data_root} contains zero parseable "
            "files. Expected at least one CSV / JSON / XLSX under "
            "<slug>-api/."
        )
    return data


def _parse_into(path: Path, mock_file: MockFile) -> None:
    suffix = mock_file.suffix

    if suffix in (".csv", ".tsv"):
        delimiter = "," if suffix == ".csv" else "\t"
        with path.open("r", encoding="utf-8", newline="") as fh:
            reader = csv.DictReader(fh, delimiter=delimiter)
            mock_file.rows = [dict(row) for row in reader]
        if not mock_file.rows:
            mock_file.raw_text = path.read_text(encoding="utf-8", errors="replace")[
                :MAX_TEXT_BYTES
            ]
        return

    if suffix == ".jsonl":
        rows: list[dict] = []
        with path.open("r", encoding="utf-8") as fh:
            for line in fh:
                line = line.strip()
                if not line:
                    continue
                rows.append(json.loads(line))
        mock_file.rows = [r for r in rows if isinstance(r, dict)]
        return

    if suffix == ".json":
        parsed = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(parsed, list):
            mock_file.rows = [r for r in parsed if isinstance(r, dict)]
        elif isinstance(parsed, dict):
            mock_file.rows = [parsed]
        return

    if suffix in (".txt", ".md"):
        mock_file.raw_text = path.read_text(encoding="utf-8", errors="replace")[
            :MAX_TEXT_BYTES
        ]
        return

    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError:
            mock_file.parse_error = (
                "openpyxl not installed — xlsx file skipped. Install "
                "openpyxl to surface cell values to the rubric prompt."
            )
            return
        wb = load_workbook(filename=str(path), data_only=True, read_only=True)
        rows: list[dict] = []
        for sheet in wb.worksheets:
            iterator = sheet.iter_rows(values_only=True)
            try:
                header = next(iterator)
            except StopIteration:
                continue
            keys = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(header)]
            for raw in iterator:
                row = {
                    keys[i]: ("" if raw[i] is None else raw[i])
                    for i in range(min(len(keys), len(raw)))
                }
                row["__sheet"] = sheet.title
                rows.append(row)
        mock_file.rows = rows
        return


def find_value(data: MockData, value: str | int | float) -> list[tuple[str, str]]:
    """Locate every mock_data file containing the given concrete value.

    Returns a list of ``(relative_path, context_snippet)`` pairs. Empty if
    the value does not appear in any active-service file — that is a signal
    to the validator that the Value Lock entry is detached from the data.
    """
    needle = str(value).strip()
    if not needle:
        return []
    hits: list[tuple[str, str]] = []

    for mock_file in data.files:
        if mock_file.rows:
            for row in mock_file.rows:
                for field_name, cell in row.items():
                    if cell is None:
                        continue
                    cell_text = str(cell)
                    if needle in cell_text:
                        snippet = f"{field_name}={cell_text}"
                        hits.append((mock_file.relative_path, snippet))
                        break
                else:
                    continue
                break
        elif mock_file.raw_text and needle in mock_file.raw_text:
            idx = mock_file.raw_text.find(needle)
            start = max(0, idx - 30)
            end = min(len(mock_file.raw_text), idx + len(needle) + 30)
            snippet = mock_file.raw_text[start:end].replace("\n", " ")
            hits.append((mock_file.relative_path, snippet))

    return hits


def sample_rows(data: MockData, max_per_file: int = 3) -> dict[str, list[dict]]:
    """Sample up to ``max_per_file`` rows per CSV/JSON file for the LLM prompt.

    The sample shows the model what real entity names / IDs look like so it
    can embed them in criteria (Rule 3 self-containment).
    """
    sample: dict[str, list[dict]] = {}
    for mock_file in data.files:
        if not mock_file.rows:
            continue
        sample[mock_file.relative_path] = mock_file.rows[:max_per_file]
    return sample


def format_mock_data_values(data: MockData, max_per_file: int = 3) -> str:
    """Render a compact LLM-prompt block showing real entities + IDs."""
    sample = sample_rows(data, max_per_file=max_per_file)
    if not sample:
        return "(no tabular mock_data rows parsed)"
    lines: list[str] = []
    for rel_path, rows in sample.items():
        lines.append(f"### `{rel_path}`")
        if not rows:
            lines.append("(empty)")
            lines.append("")
            continue
        keys = list(rows[0].keys())
        lines.append("| " + " | ".join(keys) + " |")
        lines.append("|" + "|".join(["---"] * len(keys)) + "|")
        for row in rows:
            cells = []
            for k in keys:
                v = row.get(k, "")
                cells.append(_truncate_cell(str(v) if v is not None else ""))
            lines.append("| " + " | ".join(cells) + " |")
        lines.append("")
    return "\n".join(lines)


def _truncate_cell(s: str, max_len: int = 60) -> str:
    s = s.replace("|", "\\|").replace("\n", " ").strip()
    return s if len(s) <= max_len else s[: max_len - 1] + "…"


def verify_value_lock_coverage(
    data: MockData, value_lock_entries: list
) -> tuple[list[str], list[str]]:
    """Check every Value Lock entry has at least one mock_data occurrence.

    Returns ``(present_keys, missing_keys)``. Entries whose source is
    explicitly marked "Phase-2 minted" are not required to live in
    mock_data files (they may be derived facts), but we still report
    findings for transparency.
    """
    present: list[str] = []
    missing: list[str] = []
    for entry in value_lock_entries:
        key = getattr(entry, "key", None)
        value = getattr(entry, "value", None)
        if not key or value is None:
            continue
        hits = find_value(data, value)
        if hits:
            present.append(key)
        else:
            missing.append(key)
    return present, missing
